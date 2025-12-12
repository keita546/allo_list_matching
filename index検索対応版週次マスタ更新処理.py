# -*- coding: utf-8 -*-
"""
Created by HIBI KEITA
改良版：新ロジック実装（高速化：インデックス検索対応）
"""

import pandas as pd
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from fuzzywuzzy import fuzz 
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation


def calculate_similarity(s1: str, s2: str) -> float:
    # 目的: 二つの文字列の類似度（0.0〜1.0）を計算する。
    # 理由: 商品名称（カナ）の似ている度合いを数値で測り、最適な旧品候補を見つけるため。
    if pd.isna(s1) or pd.isna(s2):
        return 0.0
    return fuzz.ratio(str(s1), str(s2)) / 100.0


def load_data(file_path: str, suffix: str) -> pd.DataFrame:
    # 目的: 指定されたファイルを読み込み、DataFrameに変換する。
    # 理由: CSV, TSV, Excelのどれでも読み込めるようにし、データの形式を統一して後続の処理でエラーが出ないようにするため。
    p = Path(file_path)
    ext = p.suffix.lower()
    
    try:
        if ext == '.csv':
            # 操作: UTF-8でCSVを読み込む。
            # 理由: 標準的なエンコーディングであるUTF-8でまず試行する。
            df = pd.read_csv(file_path, encoding='utf-8', delimiter=',', on_bad_lines='skip')
        elif ext == '.tsv':
            # 操作: TSV（タブ区切り）をUTF-8で読み込む。
            # 理由: TSVファイルにも対応するため。
            df = pd.read_csv(file_path, encoding='utf-8', delimiter='\t', on_bad_lines='skip')
        elif ext in ['.xlsx', '.xls']:
            # 操作: Excelファイルを読み込む。
            # 理由: Excel形式のデータも直接扱えるようにするため。
            df = pd.read_excel(file_path)
        else:
            raise ValueError(f"サポート外のファイル形式です:{ext}")
    except UnicodeDecodeError:
        # 操作: UTF-8で失敗した場合、Shift-JISで再試行する。
        # 理由: 日本で使われる古いCSVファイルはShift-JISでエンコードされがちなので、エラー回避のために自動で切り替える。
        if ext == '.csv':
            df = pd.read_csv(file_path, encoding='shift_jis', delimiter=',', on_bad_lines='skip')
        elif ext == '.tsv':
            df = pd.read_csv(file_path, encoding='shift_jis', delimiter='\t', on_bad_lines='skip')
        else:
            raise
    
    # 操作: データ内の「NULL」という文字列をPandasの欠損値(NA)に置き換える。
    # 理由: 欠損値のチェック（pd.isna）を正確に行えるようにするため。
    df = df.replace('NULL', pd.NA)
    
    required_cols = [
        'メーカーコード', 'ブランドコード', '標準分類コード(タイプ)',
        '目付', 'ブランド名称', '標準分類名(タイプ)',
        '商品名称（カナ）', 'JANコード', 'メーカー名称',
    ]
    
    # 操作: 必要なカラムがない場合、空のカラムを追加する。
    # 理由: 新旧マスタでカラム名が揃っていなくても、後続の処理に進めるようにするため。
    for col in required_cols:
        if col not in df.columns:
            df[col] = pd.NA 
    
    # 操作: 全てのカラム名に接尾辞（_新, _旧）を付ける。
    # 理由: 新マスタと旧マスタのデータを結合したときに、どのデータか区別できるようにするため。
    df = df.add_suffix(suffix)
    return df


def get_weight_range(weight):
    # 目的: 目付（重さ）の値から、許容される範囲（±20%）を計算する。
    # 理由: 厳密な値ではなく、少しの誤差（±20%）は許容してマッチングさせるため。
    try:
        w = float(weight)
        return w * 0.8, w * 1.2
    except (ValueError, TypeError):
        # 理由: 目付が不明な場合は、この機能ではフィルタリングしないようにするため（Noneを返す）。
        return None, None


def clean_initial_data(df: pd.DataFrame, suffix: str) -> pd.DataFrame:
    # 目的: メーカー名称、ブランドコード、標準分類コード(タイプ)の3つ全てがNULLの行を削除する。
    # 理由: マッチングに必要なキー情報が全て揃っていない行は、処理対象から外して無駄をなくすため。
    maker_col = f'メーカー名称{suffix}'
    brand_col = f'ブランドコード{suffix}'
    type_col = f'標準分類コード(タイプ){suffix}'
    
    before_count = len(df)
    # 操作: 3つのキーが全て欠損値（isna()）である行を、~（否定）で除外する。
    df_cleaned = df[
        ~(df[maker_col].isna() & df[brand_col].isna() & df[type_col].isna())
    ].copy()
    
    after_count = len(df_cleaned)
    print(f"【クレンジング{suffix}】{before_count}行 → {after_count}行 ({before_count - after_count}行削除)")
    
    return df_cleaned


def preprocess_old_data(df_old: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    # 目的: 旧マスタの事前処理と、検索高速化のためのインデックス設定を行う。
    # 理由: find_best_matchのループ内で毎回全行を検索するのではなく、
    # 検索キーをインデックスとして準備しておくことで、データの抽出速度を劇的に向上させるため。
    df_processed = df_old.copy()
    
    # 操作: 検索に使用する文字列カラムをクレンジング（前後の空白除去）し、目付を数値型に変換する。
    # 理由: マッチングの精度を上げるために、比較する文字列はきれいに揃えておく必要がある。
    df_processed['メーカー名称_旧'] = df_processed['メーカー名称_旧'].astype(str).str.strip()
    df_processed['ブランドコード_旧'] = df_processed['ブランドコード_旧'].astype(str).str.strip()
    df_processed['標準分類コード(タイプ)_旧'] = df_processed['標準分類コード(タイプ)_旧'].astype(str).str.strip()
    df_processed['目付_旧_float'] = pd.to_numeric(df_processed['目付_旧'], errors='coerce')

    # 操作: ブランドコードをインデックスとして設定したDataFrameを作成する。
    # 理由: パターン1（ブランドコード検索）を高速化するため。
    df_processed_brands_indexed = df_processed.set_index('ブランドコード_旧')
    
    # 操作: メーカー名称とタイプコードの複合インデックスを設定したDataFrameを作成する。
    # 理由: パターン2（メーカー名称+タイプコード検索）を高速化するため。
    df_processed_multi_indexed = df_processed.set_index(['メーカー名称_旧', '標準分類コード(タイプ)_旧'])

    # 理由: find_best_matchでそれぞれの検索ロジックに合わせて最適なデータフレームを使えるように、3つのDFを返す。
    return df_processed, df_processed_brands_indexed, df_processed_multi_indexed


def process_master_data(old_path: str, new_path: str, output_dir: str):
    # 目的: メインロジックとして、新品データと旧品データを突合処理し、結果を出力する。
    # 理由: アプリケーションの心臓部で、マッチング処理全体を管理するため。
    
    try:
        print("データ読み込み中...")
        df_new = load_data(new_path, '_新')
        df_old = load_data(old_path, '_旧')
        
        # 初期クレンジング
        df_new = clean_initial_data(df_new, '_新')
        df_old = clean_initial_data(df_old, '_旧')
        
        if '発売日_旧' not in df_old.columns:
            df_old['発売日_旧'] = pd.NA
        
        # 旧マスタを事前処理（高速化対応）
        print("旧マスタ前処理中...")
        df_old_processed, df_old_brands_indexed, df_old_multi_indexed = preprocess_old_data(df_old)
        
        print(f"突合処理開始...（{len(df_new)}件）")
        
        results = []
        
        # 操作: 新マスタの各行をループで回す。
        # 理由: 新品1つ1つに対して、対応する旧品を探す（マッチング）必要があるため。
        for idx, new_row in df_new.iterrows():
            if idx % 100 == 0:
                print(f"処理中... {idx}/{len(df_new)}")
            
            # 操作: find_best_matchに関数でインデックス化したDFを渡す。
            # 理由: find_best_match内で高速なインデックス検索ができるようにするため。
            result = find_best_match(new_row, df_old_processed, df_old_brands_indexed, df_old_multi_indexed, df_old)
            results.append(result)
        
        
        # 結果の整形と出力処理は省略せず記述
        print("結果を整形中...")
        analysis_result = pd.DataFrame(results)
        
        final_df = pd.concat([df_new.reset_index(drop=True), analysis_result], axis=1)
        
        final_df = final_df.rename(columns={
            '標準分類名(タイプ)_新': '標準分類(タイプ)_新',
        }, errors='ignore')
        
        if '候補あり' in final_df.columns:
            final_df = final_df[final_df['候補あり'] == True].copy()
         # 「候補あり」列を削除（出力に不要）
            final_df = final_df.drop('候補あり', axis=1)
        
        report_columns = [
            '照合結果', 'JANコード_旧','商品名称（カナ）_旧','商品名称（漢字）_旧','メーカー名称_旧','標準分類(タイプ)_旧', 
            'ブランド名称_旧','目付_旧','幅_旧','高さ_旧','奥行_旧','最高類似度','判定',
            'JANコード_新','商品名称（カナ）_新','商品名称（漢字）_新','メーカー名称_新','標準分類(タイプ)_新', 
            'ブランド名称_新','目付_新','発売日_新',
            '幅_新','高さ_新','奥行_新','候補','パターン'
        ]
        
        existing_columns = [col for col in report_columns if col in final_df.columns]
        output_df = final_df[existing_columns].copy()
        output_df = output_df.fillna('')
        
        # CSV出力
        print("CSV出力中...")
        csv_path = Path(output_dir) / 'マッチング結果.csv'
        # 理由: 処理結果をCSVファイルとして保存し、確認できるようにするため。
        output_df.to_csv(csv_path, index=False, encoding='utf-8')
        
        # 候補品リスト作成（候補ありのみ）
        print("候補品リスト作成中...")
        candidate_sheet_data = create_candidate_sheet_data(df_new, df_old_processed, df_old, results, df_old_brands_indexed, df_old_multi_indexed)
        
        # Excel出力
        print("Excel出力中...")
        excel_path = Path(output_dir) / 'マッチング結果.xlsx'
        # 理由: CSVと合わせてExcelファイルも保存し、視認性を高め、候補品リストも同時に提供するため。
        write_excel_with_dropdowns(output_df, excel_path, candidate_sheet_data)
        
        print("完了！")
        return csv_path, excel_path
    
    except Exception as e:
        # 理由: 処理中にエラーが発生した場合、それを捕捉してメッセージとして出力する。
        raise Exception(f"データ処理中にエラー: {e}")


def find_best_match(new_row: pd.Series, 
                    df_old_processed: pd.DataFrame, 
                    df_old_brands_indexed: pd.DataFrame, 
                    df_old_multi_indexed: pd.DataFrame, 
                    df_old_original: pd.DataFrame) -> dict:
    # 目的: 1行の新品データに対して、最も一致する旧品データをインデックス検索で見つける。
    # 理由: 高速に、かつ複数の条件（ブランド、タイプ、目付、名称類似度）を考慮して最適な候補を自動で選定するため。
    
    # ... (新品情報取得は省略) ...
    new_maker_name = str(new_row.get('メーカー名称_新')).strip() if pd.notna(new_row.get('メーカー名称_新')) else None
    new_brand = str(new_row.get('ブランドコード_新')).strip() if pd.notna(new_row.get('ブランドコード_新')) else None
    new_type = str(new_row.get('標準分類コード(タイプ)_新')).strip() if pd.notna(new_row.get('標準分類コード(タイプ)_新')) else None
    new_weight = new_row.get('目付_新')
    new_name = new_row.get('商品名称（カナ）_新')
    
    if new_maker_name == 'nan': new_maker_name = None
    if new_brand == 'nan': new_brand = None
    if new_type == 'nan': new_type = None
    
    skip_reasons = []
    matching_old = None
    pattern_name = '不明'  # ← ここで初期化を追加！
    
    # パターン1: ブランドあり
    if new_brand:
        pattern_name = 'ブランドのみ'  # ← デフォルトを先に設定
        # 操作: ブランドコードをキーとしてインデックス検索（df.loc[]）を行う。
        # 理由: 旧マスタ全体をチェックするのではなく、インデックスを使って一致する行だけを効率よく抽出するため。
        try:
            # df_old_brands_indexedはブランドコードがインデックスとして設定済み
            matching_old = df_old_brands_indexed.loc[new_brand].copy() 
            
            # 操作: 抽出結果が1行（Series）の場合、後の処理のためにDataFrameに変換する。
            # 理由: 目付チェックなどの処理を統一的にDataFrameとして扱えるようにするため。
            if isinstance(matching_old, pd.Series):
                 matching_old = matching_old.to_frame().T
                 
        except KeyError:
            # 理由: インデックスに一致するキーがない場合（KeyError）は、候補なしとして処理を継続する。
            matching_old = df_old_processed.iloc[0:0].copy() 

        # ... (候補なしチェックと目付チェックは変更なし) ...
        
        if matching_old.empty:
            return {
                '照合結果': '候補なし（ブランド不一致）',
                '最高類似度': 0.0, '判定': '✕', '候補': '', 'スキップ理由': '', '候補あり': False
            }
        
        # 目付チェック
        if pd.notna(new_weight):
            min_w, max_w = get_weight_range(new_weight)
            if min_w and max_w:
                weight_filtered = matching_old[
                    (matching_old['目付_旧_float'] >= min_w) & 
                    (matching_old['目付_旧_float'] <= max_w)
                ]
                if weight_filtered.empty:
                    return {
                        '照合結果': '候補なし（目付範囲外）',
                        '最高類似度': 0.0, '判定': '✕', '候補': '', 'スキップ理由': '', '候補あり': False
                    }
                matching_old = weight_filtered
                pattern_name = 'ブランド+目付'  # ← パターン名設定

        else:
            skip_reasons.append('目付スキップ')
            pattern_name = 'ブランドのみ'  # ← パターン名設定
    
    # パターン2: ブランドなし → メーカー名称+タイプ
    elif new_maker_name and new_type:
        pattern_name = 'メーカー名称+タイプのみ'  # ← デフォルトを先に設定
        # 操作: メーカー名称とタイプコードの複合キー（タプル）でインデックス検索を行う。
        # 理由: 複合キー（組み合わせ）が一致する行を効率よく抽出するため。
        try:
            # 複合インデックスのアクセスにはタプル (キー1, キー2) を使用
            matching_old = df_old_multi_indexed.loc[(new_maker_name, new_type)].copy()
            
            if isinstance(matching_old, pd.Series):
                 matching_old = matching_old.to_frame().T
                 
        except KeyError:
             # 理由: インデックスに一致する複合キーがない場合、候補なしとして処理を継続する。
             matching_old = df_old_processed.iloc[0:0].copy()
            
        
        # ... (候補なしチェックと目付チェックは変更なし) ...
        
        if matching_old.empty:
            return {
                '照合結果': '候補なし（メーカー名称+タイプ不一致）',
                '最高類似度': 0.0, '判定': '✕', '候補': '', 'スキップ理由': '', '候補あり': False
            }
        
        # 目付チェック
        if pd.notna(new_weight):
            min_w, max_w = get_weight_range(new_weight)
            if min_w and max_w:
                weight_filtered = matching_old[
                    (matching_old['目付_旧_float'] >= min_w) & 
                    (matching_old['目付_旧_float'] <= max_w)
                ]
                if weight_filtered.empty:
                    return {
                        '照合結果': '候補なし（目付範囲外）',
                        '最高類似度': 0.0, '判定': '✕', '候補': '', 'スキップ理由': '', '候補あり': False
                    }
                matching_old = weight_filtered
                pattern_name = 'メーカー名称+タイプ+目付'  # ← パターン名設定
        else:
            skip_reasons.append('目付スキップ')
            pattern_name = 'メーカー名称+タイプのみ'  # ← パターン名設定
    
    else:
        # 理由: マッチングに必要なキーコード（ブランド、またはメーカー+タイプ）が不足している場合、候補なしとする。
        return {
            '照合結果': '候補なし（キーコード不足）',
            '最高類似度': 0.0, '判定': '✕', '候補': '', 'スキップ理由': '', '候補あり': False
        }
    
    # ... (名称一致で最高類似度を選択する処理は変更なし) ...
    # 名称一致で最高類似度を選択
    candidates = matching_old[['商品名称（カナ）_旧', 'JANコード_旧']].drop_duplicates()
    
    if candidates.empty:
        return {
            '照合結果': '候補なし（名称一致なし）',
            '最高類似度': 0.0, '判定': '✕', '候補': '',
            'スキップ理由': '、'.join(skip_reasons) if skip_reasons else '', '候補あり': False
        }
    
    similarities = [
        (calculate_similarity(new_name, row['商品名称（カナ）_旧']), 
         row['商品名称（カナ）_旧'], 
         row['JANコード_旧'])
        for _, row in candidates.iterrows()
    ]
    similarities.sort(key=lambda x: x[0], reverse=True)
    
    best_score, best_name, best_jan = similarities[0]
    best_old_row = df_old_original[df_old_original['JANコード_旧'] == best_jan].iloc[0]
    
    # 判定
    if best_score >= 0.8:
        result = '高類似度候補あり (80%以上)'
        judgment = '○'
    else:
        result = '低類似度 (80%未満・要手動確認)'
        judgment = '✕'

    return {
        '照合結果': result,
        '最高類似度': best_score,
        '判定': judgment,
        '候補': f"{best_name}({best_score:.1%})",
        'スキップ理由': '、'.join(skip_reasons) if skip_reasons else '',
        '候補あり': True,
        'JANコード_旧': best_old_row.get('JANコード_旧', ''),
        '商品名称（カナ）_旧': best_old_row.get('商品名称（カナ）_旧', ''),
        'メーカー名称_旧': best_old_row.get('メーカー名称_旧', ''),
        '標準分類(タイプ)_旧': best_old_row.get('標準分類名(タイプ)_旧', ''),
        'ブランド名称_旧': best_old_row.get('ブランド名称_旧', ''),
        '目付_旧': best_old_row.get('目付_旧', ''),
        '発売日_旧': best_old_row.get('発売日_旧', ''),
        '商品名称（漢字）_旧': best_old_row.get('商品名称（漢字）_旧', ''),
        '幅_旧': best_old_row.get('幅_旧', ''),
        '高さ_旧': best_old_row.get('高さ_旧', ''),
        '奥行_旧': best_old_row.get('奥行_旧', ''),
        'パターン': pattern_name
    }


def create_candidate_sheet_data(df_new, df_old_processed, df_old_original, results, df_old_brands_indexed, df_old_multi_indexed) -> list:
    # 目的: 候補ありと判定された新品に対して、すべてのマッチングパターンで抽出された候補品リストを作成する。
    # 理由: メインのマッチングで選ばれた候補以外に、手動で確認できるように「次点」の候補をまとめてExcelシートに出力するため。
    candidates_list = []
    
    for idx, new_row in df_new.iterrows():
        # ... (初期設定と候補なしスキップは省略) ...        
        if idx % 100 == 0:
            print(f"候補品処理中... {idx}/{len(df_new)}")
        
        if not results[idx].get('候補あり', False):
            continue
        
        new_maker_name = str(new_row.get('メーカー名称_新')).strip() if pd.notna(new_row.get('メーカー名称_新')) else None
        new_brand = str(new_row.get('ブランドコード_新')).strip() if pd.notna(new_row.get('ブランドコード_新')) else None
        new_type = str(new_row.get('標準分類コード(タイプ)_新')).strip() if pd.notna(new_row.get('標準分類コード(タイプ)_新')) else None
        new_name = new_row.get('商品名称（カナ）_新', '')
        new_kanji = new_row.get('商品名称（漢字）_新', '')
        new_weight = new_row.get('目付_新')
        
        if new_maker_name == 'nan': new_maker_name = None
        if new_brand == 'nan': new_brand = None
        if new_type == 'nan': new_type = None
        
        new_jan = new_row.get('JANコード_新', '')
        new_maker_name_val = new_row.get('メーカー名称_新', '')
        new_type_name = new_row.get('標準分類名(タイプ)_新', '')  # タイプに変更
        new_brand_name = new_row.get('ブランド名称_新', '')
        new_wide = new_row.get('幅_新', '')
        new_height = new_row.get('高さ_新', '')
        new_length = new_row.get('奥行_新', '')
        new_release = new_row.get('発売日_新','')
        
        matched_df = None
        pattern_name = ''

        # 【旧品検索処理の高速化】: ブールインデックス検索（全行チェック）ではなく、インデックス検索（df.loc）を使用する。
        
        # ===== 優先順位付きパターン選択 =====
        
        # パターンA: ブランド+目付（最優先）
        if new_brand and pd.notna(new_weight):
            min_w, max_w = get_weight_range(new_weight)
            if min_w and max_w:
                try:
                    brand_matched = df_old_brands_indexed.loc[new_brand].copy()
                    if isinstance(brand_matched, pd.Series):
                        brand_matched = brand_matched.to_frame().T
                    
                    pattern_a = brand_matched[
                        (brand_matched['目付_旧_float'] >= min_w) &
                        (brand_matched['目付_旧_float'] <= max_w)
                    ].copy()
                    
                    if not pattern_a.empty:
                        matched_df = pattern_a
                        pattern_name = 'ブランド+目付'
                except KeyError:
                    pass
        
        # パターンB: ブランドのみ（Aがなければ）
        if matched_df is None and new_brand:
            try:
                pattern_b = df_old_brands_indexed.loc[new_brand].copy()
                if isinstance(pattern_b, pd.Series):
                    pattern_b = pattern_b.to_frame().T
                
                if not pattern_b.empty:
                    matched_df = pattern_b
                    pattern_name = 'ブランドのみ'
            except KeyError:
                pass
        
        # パターンC: メーカー名称+タイプ+目付（A, Bがなければ）
        if matched_df is None and new_maker_name and new_type and pd.notna(new_weight):
            min_w, max_w = get_weight_range(new_weight)
            if min_w and max_w:
                try:
                    multi_matched = df_old_multi_indexed.loc[(new_maker_name, new_type)].copy()
                    if isinstance(multi_matched, pd.Series):
                        multi_matched = multi_matched.to_frame().T
                    
                    pattern_c = multi_matched[
                        (multi_matched['目付_旧_float'] >= min_w) &
                        (multi_matched['目付_旧_float'] <= max_w)
                    ].copy()
                    
                    if not pattern_c.empty:
                        matched_df = pattern_c
                        pattern_name = 'メーカー名称+タイプ+目付'
                except KeyError:
                    pass
        
        # パターンD: メーカー名称+タイプのみ（A, B, Cがなければ）
        if matched_df is None and new_maker_name and new_type:
            try:
                pattern_d = df_old_multi_indexed.loc[(new_maker_name, new_type)].copy()
                if isinstance(pattern_d, pd.Series):
                    pattern_d = pattern_d.to_frame().T
                
                if not pattern_d.empty:
                    matched_df = pattern_d
                    pattern_name = 'メーカー名称+タイプのみ'
            except KeyError:
                pass
        
        # ===== 候補品リスト作成 =====
        if matched_df is not None and not matched_df.empty:
            matched_df = matched_df.drop_duplicates(subset=['JANコード_旧'])
            
            # メーカー名称を補完
            if 'メーカー名称_旧' not in matched_df.columns:
                maker_info = df_old_original[['JANコード_旧', 'メーカー名称_旧']].drop_duplicates()
                matched_df = matched_df.merge(maker_info, on='JANコード_旧', how='left')
            
            # 類似度計算
            similarities = [
                (calculate_similarity(new_name, row.get('商品名称（カナ）_旧', '')), row)
                for _, row in matched_df.iterrows()
            ]
            similarities.sort(key=lambda x: x[0], reverse=True)
            
            for score, old_row in similarities:
                candidates_list.append({
                    '旧候補JANコード': old_row.get('JANコード_旧', ''),
                    '旧候補商品名': old_row.get('商品名称（カナ）_旧', ''),
                    '旧商品名称（漢字）': old_row.get('商品名称（漢字）_旧', ''),
                    '旧メーカー名称': old_row.get('メーカー名称_旧', ''),
                    '旧標準分類名(タイプ)': old_row.get('標準分類名(タイプ)_旧', ''),  # タイプに変更
                    '旧ブランド名称': old_row.get('ブランド名称_旧', ''),
                    '旧目付': old_row.get('目付_旧', ''),
                    '旧幅': old_row.get('幅_旧', ''),
                    '旧高さ': old_row.get('高さ_旧', ''),
                    '旧奥行': old_row.get('奥行_旧', ''),
                    '類似度': f'{score:.1%}',
                    '新JANコード': new_jan,
                    '新商品名': new_name,
                    '新商品名漢字': new_kanji,
                    '新メーカー名称': new_maker_name_val,
                    '新標準分類名(タイプ)': new_type_name,  # タイプに変更
                    '新ブランド名称': new_brand_name,
                    '新目付': new_weight,
                    '新発売日': new_release,
                    '新幅': new_wide,
                    '新高さ': new_height,
                    '新奥行': new_length,
                    'パターン': pattern_name,
                })
    
    return candidates_list


# ... (write_excel_with_dropdowns, MasterMatcherApp, main処理は変更なし) ...

def write_excel_with_dropdowns(df, output_path, candidate_data=None):
    """Excel出力"""
    wb = Workbook()
    ws = wb.active
    ws.title = "マッチング"
    
    # ヘッダー
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # データ
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # 列幅調整
    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20
    
    # 候補品リスト
    if candidate_data:
        df_candidates = pd.DataFrame(candidate_data).fillna('')
        ws_candidates = wb.create_sheet("候補品リスト")
        
        for col_idx, col_name in enumerate(df_candidates.columns, 1):
            cell = ws_candidates.cell(row=1, column=col_idx, value=col_name)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for row_idx, row in enumerate(dataframe_to_rows(df_candidates, index=False, header=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_candidates.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        for col_idx in range(1, len(df_candidates.columns) + 1):
            ws_candidates.column_dimensions[ws_candidates.cell(row=1, column=col_idx).column_letter].width = 20
    
    # 理由: Excelファイルを指定されたパスに保存する。
    wb.save(output_path)


class MasterMatcherApp:
    def __init__(self, master):
        self.master = master
        master.title("旧品マッチング（新→旧方向）")
        master.geometry("400x250")
        
        self.old_path_var = tk.StringVar()
        self.new_path_var = tk.StringVar()
        # 理由: デスクトップをデフォルトの出力先とする。
        self.output_dir_var = tk.StringVar(value=os.path.join(os.path.expanduser('~'), 'Desktop'))
        
        self.create_widgets(master)
    
    def create_widgets(self, master):
        main_frame = tk.Frame(master, padx=10, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(main_frame, text="旧マスタファイル (.csv, .tsv, .xlsx)").grid(row=0, column=0, sticky='w', pady=5)
        tk.Entry(main_frame, textvariable=self.old_path_var, width=50).grid(row=1, column=0, padx=5, sticky='ew')
        tk.Button(main_frame, text="参照", command=self.select_file_old).grid(row=1, column=1, padx=5)
        
        tk.Label(main_frame, text="新マスタファイル (.csv, .tsv, .xlsx)").grid(row=2, column=0, sticky='w', pady=5)
        tk.Entry(main_frame, textvariable=self.new_path_var, width=50).grid(row=3, column=0, padx=5, sticky='ew')
        tk.Button(main_frame, text="参照", command=self.select_file_new).grid(row=3, column=1, padx=5)
        
        tk.Label(main_frame, text="結果出力先フォルダ").grid(row=4, column=0, sticky='w', pady=5)
        tk.Entry(main_frame, textvariable=self.output_dir_var, width=50).grid(row=5, column=0, padx=5, sticky='ew')
        tk.Button(main_frame, text="参照", command=self.select_output_dir).grid(row=5, column=1, padx=5)
        
        tk.Button(main_frame, text="マッチング開始", command=self.execute_analysis,
                  bg="blue", fg="white", font=('Helvetica', 12, 'bold')).grid(row=6, column=0, columnspan=2, pady=20, sticky='ew')
    
    def select_file(self, path_var):
        file_types = [
            ("マスタファイル", "*.csv *.tsv *.xlsx *.xls"),
            ("CSVファイル", "*.csv"),
            ("TSVファイル", "*.tsv"),
            ("Excelファイル", "*.xlsx *.xls")
        ]
        file_path = filedialog.askopenfilename(parent=self.master, title="ファイルを選択", filetypes=file_types)
        if file_path:
            path_var.set(file_path)
    
    def select_file_old(self):
        self.select_file(self.old_path_var)
    
    def select_file_new(self):
        self.select_file(self.new_path_var)
    
    def select_output_dir(self):
        folder_path = filedialog.askdirectory(parent=self.master, title="保存フォルダを選択")
        if folder_path:
            self.output_dir_var.set(folder_path)
    
    def execute_analysis(self):
        new_path = self.new_path_var.get()
        old_path = self.old_path_var.get()
        output_dir = self.output_dir_var.get()
        
        if not all([new_path, old_path, output_dir]):
            messagebox.showwarning("入力不足", "すべて選択してください")
            return
        
        try:
            # 理由: メインの処理関数を呼び出し、結果パスを受け取る。
            csv_path, excel_path = process_master_data(old_path, new_path, output_dir)
            # 理由: 処理が成功した場合、完了メッセージと出力先を表示する。
            messagebox.showinfo("完了", f"✅ 完了\n\nCSV: {csv_path}\nExcel: {excel_path}")
        except Exception as e:
            # 理由: エラーが発生した場合、エラーメッセージを表示してユーザーに通知する。
            messagebox.showerror("エラー", f"❌ エラー発生\n{e}")


if __name__ == "__main__":
    # 理由: GUIアプリケーションのルートウィンドウを作成し、実行する。
    root = tk.Tk()
    app = MasterMatcherApp(root)
    root.mainloop()