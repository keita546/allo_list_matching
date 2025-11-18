# -*- coding: utf-8 -*-
"""
Created by HIBI KEITA
修正版：メーカー・ブランド・タイプ絶対一致 → 目付幅判定 → 類似度判定
"""

import pandas as pd
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
from fuzzywuzzy import fuzz 
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows


def calculate_similarity(s1: str, s2: str) -> float:
    """二つの文字列の類似度（0.0〜1.0）を計算します。"""
    return fuzz.ratio(s1, s2) / 100.0


def load_data(file_path: str, suffix: str) -> pd.DataFrame:
    """ファイルを読み込み、suffixをカラムに付与"""
    p = Path(file_path)
    ext = p.suffix.lower()
    
    try:
        if ext == '.csv':
            df = pd.read_csv(file_path, encoding='utf-8', delimiter=',', on_bad_lines='skip')
        elif ext == '.tsv':
            df = pd.read_csv(file_path, encoding='utf-8', delimiter='\t', on_bad_lines='skip')
        elif ext in ['.xlsx', '.xls']:
            df = pd.read_excel(file_path)
        else:
            raise ValueError(f"サポート外のファイル形式です:{ext}")
    except UnicodeDecodeError:
        if ext == '.csv':
            df = pd.read_csv(file_path, encoding='shift_jis', delimiter=',', on_bad_lines='skip')
        elif ext == '.tsv':
            df = pd.read_csv(file_path, encoding='shift_jis', delimiter='\t', on_bad_lines='skip')
        else:
            raise
    
    df = df.replace('NULL', pd.NA)

    required_cols = [
        'メーカーコード', 
        'ブランドコード',
        '標準分類コード(タイプ)',
        '目付', 
        'ブランド名称', 
        '標準分類名(クラス)',
        '商品名称（カナ）',
        'JANコード',
    ]
    
    for col in required_cols:
        if col not in df.columns:
            df[col] = pd.NA 

    df = df.add_suffix(suffix)
    return df


def get_weight_range(weight):
    """目付から許容範囲を計算（100~90%の幅）"""
    try:
        w = float(weight)
        min_weight = w * 0.9
        return min_weight, w
    except (ValueError, TypeError):
        return None, None


def process_master_data(old_path: str, new_path: str, output_dir: str):
    """旧マスタと新マスタを突合し、リニューアル品を抽出するメインロジック関数"""
    
    try:
        df_old = load_data(old_path, '_旧')
        df_new = load_data(new_path, '_新')
        
        # デバッグ：列名を出力
        debug_cols_path = Path(output_dir) / 'debug_列名.txt'
        with open(debug_cols_path, 'w', encoding='utf-8') as f:
            f.write("旧マスタの列名:\n")
            f.write(str(df_old.columns.tolist()) + "\n\n")
            f.write("新マスタの列名:\n")
            f.write(str(df_new.columns.tolist()) + "\n\n")
            f.write("旧マスタの最初の行:\n")
            f.write(str(df_old.iloc[0].to_dict()) + "\n")
        
        if '発売日_新' not in df_new.columns:
            df_new['発売日_新'] = pd.NA

        def find_best_matches_for_row(old_row: pd.Series) -> pd.Series:
            
            old_maker_code = str(old_row.get('メーカーコード_旧')).strip() if pd.notna(old_row.get('メーカーコード_旧')) else None
            old_brand_code = str(old_row.get('ブランドコード_旧')).strip() if pd.notna(old_row.get('ブランドコード_旧')) else None
            old_type_code = str(old_row.get('標準分類コード(タイプ)_旧')).strip() if pd.notna(old_row.get('標準分類コード(タイプ)_旧')) else None
            old_weight = old_row.get('目付_旧')
            old_product_name_kana = old_row.get('商品名称（カナ）_旧')
            
            # 3つのコードが空の場合
            if old_maker_code is None or old_brand_code is None or old_type_code is None:
                return pd.Series({
                    '照合結果': '候補なし（キーコード空）', 
                    '最高類似度': 0.0,
                    '判定': '✕',
                    '候補': '',
                    '候補_全リスト': '',
                })

            # 新マスタのコード列を文字列に統一
            df_new_match = df_new.copy()
            df_new_match['メーカーコード_新'] = df_new_match['メーカーコード_新'].astype(str).str.strip()
            df_new_match['ブランドコード_新'] = df_new_match['ブランドコード_新'].astype(str).str.strip()
            df_new_match['標準分類コード(タイプ)_新'] = df_new_match['標準分類コード(タイプ)_新'].astype(str).str.strip()
            
            # ステップ1: メーカー・ブランド・タイプコードで絶対一致
            matching_new = df_new_match[
                (df_new_match['メーカーコード_新'] == old_maker_code) &
                (df_new_match['ブランドコード_新'] == old_brand_code) &
                (df_new_match['標準分類コード(タイプ)_新'] == old_type_code)
            ]
            
            if matching_new.empty:
                return pd.Series({
                    '照合結果': '候補なし（3コード不一致）', 
                    '最高類似度': 0.0,
                    '判定': '✕',
                    '候補': '',
                    '候補_全リスト': '',
                })
            
            # ステップ2: 目付で幅判定（旧品の90%~100%）
            min_w, max_w = get_weight_range(old_weight)
            
            if min_w is not None and max_w is not None:
                # 新品の目付も数値化
                weight_filtered = matching_new.copy()
                weight_filtered['目付_新_float'] = pd.to_numeric(weight_filtered['目付_新'], errors='coerce')
                
                # 範囲内の商品を抽出（0や空白も含める）
                weight_filtered = weight_filtered[
                    (weight_filtered['目付_新_float'] >= min_w) & 
                    (weight_filtered['目付_新_float'] <= max_w)
                ]
                weight_filtered = weight_filtered.drop('目付_新_float', axis=1)
            else:
                weight_filtered = matching_new
            
            if weight_filtered.empty:
                return pd.Series({
                    '照合結果': '候補なし（目付範囲外）', 
                    '最高類似度': 0.0,
                    '判定': '✕',
                    '候補': '',
                    '候補_全リスト': '',
                })
            
            # ステップ3: 類似度計算とソート
            new_candidates = weight_filtered[['商品名称（カナ）_新', 'JANコード_新']].drop_duplicates()
            
            similarities = []
            for _, row in new_candidates.iterrows():
                if pd.notna(old_product_name_kana):
                    score = calculate_similarity(str(old_product_name_kana), str(row['商品名称（カナ）_新']))
                else:
                    score = 0.0
                similarities.append((score, row['商品名称（カナ）_新'], row['JANコード_新']))
            
            similarities.sort(key=lambda x: x[0], reverse=True)
            
            # 最高類似度の新品を取得
            best_score, best_name, best_jan = similarities[0]
            best_new_row = df_new[df_new['JANコード_新'] == best_jan].iloc[0]

            # 新品情報を抽出
            new_info_cols = [col for col in ['JANコード_新', '商品名称（カナ）_新', 'メーカー名称_新', 
                                              '標準分類名(クラス)_新', 'ブランド名称_新', '目付_新', '発売日_新'] 
                            if col in best_new_row.index]
            new_info_for_report = best_new_row[new_info_cols]
            
            # カラム名をリネーム
            new_info_for_report = new_info_for_report.rename(
                {'標準分類名(クラス)_新': '標準分類(クラス)_新'},
                errors='ignore'
            )
            
            # 全候補を類似度が高い順にリスト化（ドロップダウン用、改行区切り）
            all_candidates_str = "\n".join([f"{name}({score:.1%})" for score, name, _ in similarities])
            
            # 照合結果と判定
            if best_score >= 0.8: 
                result = '高類似度候補あり (80%以上)'
                judgment = '○'
                candidate_str = f"{best_name}({best_score:.1%})"
            else:
                result = '低類似度 (80%未満・要手動確認)'
                judgment = '✕'
                candidate_str = f"{best_name}({best_score:.1%})"

            result_series = pd.Series({
                '照合結果': result, 
                '最高類似度': best_score,
                '判定': judgment,
                '候補': candidate_str,
                '候補_全リスト': all_candidates_str,
            })
            return pd.concat([result_series, new_info_for_report])

        # 分析を実行
        analysis_result = df_old.apply(find_best_matches_for_row, axis=1)

        # 旧品DFと分析結果を結合
        final_df = pd.concat([df_old, analysis_result], axis=1)

        # カラム名をリネーム
        final_df = final_df.rename(columns={
            '標準分類名(クラス)_旧': '標準分類(クラス)_旧',
        }, errors='ignore')
        
        # 出力カラムを定義
        report_columns = [
            '照合結果', 'JANコード_旧','商品名称（カナ）_旧','メーカー名称_旧','標準分類(クラス)_旧', 'ブランド名称_旧','目付_旧','最高類似度','判定',
            'JANコード_新','商品名称（カナ）_新','メーカー名称_新','標準分類(クラス)_新', 'ブランド名称_新','目付_新','発売日_新',
            '候補', '候補_全リスト',
        ]
        
        # 存在するカラムのみを出力
        existing_columns = [col for col in report_columns if col in final_df.columns]
        output_df = final_df[existing_columns].copy()
        
        # NA値を空文字に置換（Excel出力エラー対策）
        output_df = output_df.fillna('')
        
        # CSV出力
        csv_path = Path(output_dir) / 'リニューアル品抽出結果.csv'
        output_df.to_csv(csv_path, index=False, encoding='utf-8')
        
        # 候補品シート用データを別途作成
        candidate_sheet_data = create_candidate_sheet_data(df_old, df_new)
        
        # デバッグ：候補品データをCSVに出力
        debug_path = Path(output_dir) / 'debug_候補品データ.csv'
        debug_info = {
            'メッセージ': [
                f'total rows: {len(final_df)}',
                f'candidate_sheet_data length: {len(candidate_sheet_data)}',
                f'照合結果の種類: {final_df["照合結果"].unique().tolist() if "照合結果" in final_df.columns else "N/A"}',
                f'候補_全リストに値がある件数: {final_df["候補_全リスト"].notna().sum() if "候補_全リスト" in final_df.columns else 0}',
            ]
        }
        pd.DataFrame(debug_info).to_csv(debug_path, index=False, encoding='utf-8', errors='ignore')
        
        # Excel出力（ドロップダウン機能付き + 候補品シート）
        excel_path = Path(output_dir) / 'リニューアル品抽出結果.xlsx'
        write_excel_with_dropdowns(output_df, excel_path, candidate_sheet_data)
        
        return csv_path, excel_path

    except Exception as e:
        raise Exception(f"データ処理中にエラーが発生しました。詳細: {e}")


def create_candidate_sheet_data(df_old, df_new) -> list:
    """候補品データを別シート用に構築（3キー絶対一致したもの全部を類似度降順で出力）"""
    candidates_list = []
    
    for _, old_row in df_old.iterrows():
        old_maker_code = str(old_row.get('メーカーコード_旧')).strip() if pd.notna(old_row.get('メーカーコード_旧')) else None
        old_brand_code = str(old_row.get('ブランドコード_旧')).strip() if pd.notna(old_row.get('ブランドコード_旧')) else None
        old_type_code = str(old_row.get('標準分類コード(タイプ)_旧')).strip() if pd.notna(old_row.get('標準分類コード(タイプ)_旧')) else None
        old_product_name_kana = old_row.get('商品名称（カナ）_旧', '')
        
        old_jan = old_row.get('JANコード_旧', '')
        old_name = old_row.get('商品名称（カナ）_旧', '')
        old_maker = old_row.get('メーカー名称_旧', '')
        old_class = old_row.get('標準分類(クラス)_旧', '')
        old_brand = old_row.get('ブランド名称_旧', '')
        old_weight = old_row.get('目付_旧', '')
        
        # 3キー絶対一致した新品を取得
        if old_maker_code and old_brand_code and old_type_code:
            df_new_match = df_new.copy()
            df_new_match['メーカーコード_新'] = df_new_match['メーカーコード_新'].astype(str).str.strip()
            df_new_match['ブランドコード_新'] = df_new_match['ブランドコード_新'].astype(str).str.strip()
            df_new_match['標準分類コード(タイプ)_新'] = df_new_match['標準分類コード(タイプ)_新'].astype(str).str.strip()
            
            matching_new = df_new_match[
                (df_new_match['メーカーコード_新'] == old_maker_code) &
                (df_new_match['ブランドコード_新'] == old_brand_code) &
                (df_new_match['標準分類コード(タイプ)_新'] == old_type_code)
            ]
            
            # 類似度を計算して降順でソート
            similarities = []
            for _, new_row in matching_new.iterrows():
                new_name = new_row.get('商品名称（カナ）_新', '')
                if pd.notna(old_product_name_kana) and pd.notna(new_name):
                    score = calculate_similarity(str(old_product_name_kana), str(new_name))
                else:
                    score = 0.0
                similarities.append((score, new_row))
            
            # 類似度で降順ソート
            similarities.sort(key=lambda x: x[0], reverse=True)
            
            # 候補リストに追加
            for score, new_row in similarities:
                candidates_list.append({
                    '旧JANコード': old_jan,
                    '旧商品名': old_name,
                    '旧メーカー名称': old_maker,
                    '旧標準分類名(クラス)': old_class,
                    '旧ブランド名称': old_brand,
                    '旧目付': old_weight,
                    '新候補JANコード': new_row.get('JANコード_新', ''),
                    '新候補商品名': new_row.get('商品名称（カナ）_新', ''),
                    '新メーカー名称': new_row.get('メーカー名称_新', ''),
                    '新標準分類名(クラス)': new_row.get('標準分類名(クラス)_新', ''),
                    '新ブランド名称': new_row.get('ブランド名称_新', ''),
                    '新目付': new_row.get('目付_新', ''),
                    '新発売日': new_row.get('発売日_新', ''),
                    '類似度': f'{score:.1%}',
                })
    
    return candidates_list


def write_excel_with_dropdowns(df, output_path, candidate_data=None):
    """ドロップダウン機能付きExcelファイルを生成（複数シート対応）"""
    wb = Workbook()
    
    # Sheet1: 通常の照合結果
    ws = wb.active
    ws.title = "リニューアル品"
    
    # ヘッダー行を書き込み
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # データを書き込み
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # 「候補_全リスト」列にドロップダウンを追加
    if '候補_全リスト' in df.columns:
        candidate_col = df.columns.get_loc('候補_全リスト') + 1
        for row_idx in range(2, len(df) + 2):
            candidate_list = df.iloc[row_idx - 2]['候補_全リスト']
            if pd.notna(candidate_list) and candidate_list:
                # 改行で区切られた候補をドロップダウン化
                candidates = candidate_list.split('\n')
                # Excelの制限に対応（改行文字を使用）
                formula = '"' + ','.join(candidates) + '"'
                dv_candidate = DataValidation(type="list", formula1=formula, allow_blank=True)
                dv_candidate.error = '候補から選択してください'
                dv_candidate.errorTitle = '入力エラー'
                ws.add_data_validation(dv_candidate)
                dv_candidate.add(ws.cell(row=row_idx, column=candidate_col))
    
    # 列幅を自動調整
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20
    
    # Sheet2: 候補品リスト
    if candidate_data:
        df_candidates = pd.DataFrame(candidate_data).fillna('')
        ws_candidates = wb.create_sheet("候補品リスト")
        
        # ヘッダー行を書き込み
        for col_idx, col_name in enumerate(df_candidates.columns, 1):
            cell = ws_candidates.cell(row=1, column=col_idx, value=col_name)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # データを書き込み
        for row_idx, row in enumerate(dataframe_to_rows(df_candidates, index=False, header=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_candidates.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # 列幅を自動調整
        for col_idx, col_name in enumerate(df_candidates.columns, 1):
            ws_candidates.column_dimensions[ws_candidates.cell(row=1, column=col_idx).column_letter].width = 20
    
    wb.save(output_path)


class MasterMatcherApp:
    
    def __init__(self, master):
        self.master = master
        master.title("商品リニューアル品抽出")
        master.geometry("400x250")
        
        self.old_path_var = tk.StringVar()
        self.new_path_var = tk.StringVar()
        self.output_dir_var = tk.StringVar(value=os.path.join(os.path.expanduser('~'), 'Desktop'))
        
        self.create_widgets(master)

    def create_widgets(self, master):
        main_frame = tk.Frame(master, padx=10, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(main_frame, text="旧マスタファイル (.csv, .tsv, .xlsx)").grid(row=0, column=0, sticky='w', pady=5)
        tk.Entry(main_frame, textvariable=self.old_path_var, width=50).grid(row=1, column=0, padx=5, sticky='ew')
        tk.Button(main_frame, text="参照", command=self.select_file_old).grid(row=1, column=1, padx=5)

        tk.Label(main_frame, text="新マスタファイル (.csv, .tsv, .xlsx)").grid(row=2, column=0, sticky='w', pady=5)
        tk.Entry(main_frame, textvariable=self.new_path_var, width=50).grid(row=3, column=0, padx=5, sticky='ew')
        tk.Button(main_frame, text="参照", command=self.select_file_new).grid(row=3, column=1, padx=5)

        tk.Label(main_frame, text="結果出力先フォルダ").grid(row=4, column=0, sticky='w', pady=5)
        tk.Entry(main_frame, textvariable=self.output_dir_var, width=50).grid(row=5, column=0, padx=5, sticky='ew')
        tk.Button(main_frame, text="参照", command=self.select_output_dir).grid(row=5, column=1, padx=5)

        tk.Button(main_frame, text="リニューアル品抽出を実行！", command=self.execute_analysis, 
                  bg="skyblue", fg="white", font=('Helvetica', 12, 'bold')).grid(row=6, column=0, columnspan=2, pady=20, sticky='ew')

    def select_file(self, path_var: tk.StringVar):
        file_types = [
            ("マスタファイル", "*.csv *.tsv *.xlsx *.xls"),
            ("CSVファイル", "*.csv"),
            ("TSVファイル", "*.tsv"),
            ("Excelファイル", "*.xlsx *.xls")
        ]
        
        file_path = filedialog.askopenfilename(
            parent=self.master, 
            title="ファイルを選択してください", 
            filetypes=file_types
        )
        if file_path:
            path_var.set(file_path)

    def select_file_old(self):
        self.select_file(self.old_path_var)

    def select_file_new(self):
        self.select_file(self.new_path_var)

    def select_output_dir(self):
        folder_path = filedialog.askdirectory(parent=self.master, title="結果を保存するフォルダを選択してください")
        if folder_path:
            self.output_dir_var.set(folder_path)

    def execute_analysis(self):
        old_path = self.old_path_var.get()
        new_path = self.new_path_var.get()
        output_dir = self.output_dir_var.get()
        
        if not all([old_path, new_path, output_dir]):
            messagebox.showwarning("入力不足", "旧マスタ、新マスタ、出力先フォルダをすべて選択してください。")
            return

        try:
            csv_path, excel_path = process_master_data(old_path, new_path, output_dir)
            messagebox.showinfo("完了", f"✅ リニューアル品抽出が完了しました。\n\nCSV: {csv_path}\nExcel: {excel_path}")

        except Exception as e:
            messagebox.showerror("エラー", f"❌ 処理中にエラーが発生しました。\n詳細: {e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = MasterMatcherApp(root)
    root.mainloop()