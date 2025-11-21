# -*- coding: utf-8 -*-
"""
Created by HIBI KEITA
改良版：新ロジック実装
"""

import pandas as pd
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from fuzzywuzzy import fuzz 
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation


def calculate_similarity(s1: str, s2: str) -> float:
    """二つの文字列の類似度（0.0〜1.0）を計算"""
    if pd.isna(s1) or pd.isna(s2):
        return 0.0
    return fuzz.ratio(str(s1), str(s2)) / 100.0


def load_data(file_path: str, suffix: str) -> pd.DataFrame:
    """ファイルを読み込み、suffixをカラムに付与"""
    p = Path(file_path)
    ext = p.suffix.lower()
    
    try:
        if ext == '.csv':
            df = pd.read_csv(file_path, encoding='utf-8', delimiter=',', on_bad_lines='skip')
        elif ext == '.tsv':
            df = pd.read_csv(file_path, encoding='utf-8', delimiter='\t', on_bad_lines='skip')
        elif ext in ['.xlsx', '.xls']:
            df = pd.read_excel(file_path)
        else:
            raise ValueError(f"サポート外のファイル形式です:{ext}")
    except UnicodeDecodeError:
        if ext == '.csv':
            df = pd.read_csv(file_path, encoding='shift_jis', delimiter=',', on_bad_lines='skip')
        elif ext == '.tsv':
            df = pd.read_csv(file_path, encoding='shift_jis', delimiter='\t', on_bad_lines='skip')
        else:
            raise
    
    df = df.replace('NULL', pd.NA)
    
    required_cols = [
        'メーカーコード', 'ブランドコード', '標準分類コード(タイプ)',
        '目付', 'ブランド名称', '標準分類名(クラス)',
        '商品名称（カナ）', 'JANコード', 'メーカー名称',
    ]
    
    for col in required_cols:
        if col not in df.columns:
            df[col] = pd.NA 
    
    df = df.add_suffix(suffix)
    return df


def get_weight_range(weight):
    """目付から許容範囲を計算（90%~110%）"""
    try:
        w = float(weight)
        return w * 0.9, w * 1.1
    except (ValueError, TypeError):
        return None, None


def clean_initial_data(df: pd.DataFrame, suffix: str) -> pd.DataFrame:
    """初期データクレンジング：3コード全てNULLの行を削除"""
    maker_col = f'メーカー名称{suffix}'
    brand_col = f'ブランドコード{suffix}'
    type_col = f'標準分類コード(タイプ){suffix}'
    
    before_count = len(df)
    df_cleaned = df[
        ~(df[maker_col].isna() & df[brand_col].isna() & df[type_col].isna())
    ].copy()
    
    after_count = len(df_cleaned)
    print(f"【クレンジング{suffix}】{before_count}行 → {after_count}行 ({before_count - after_count}行削除)")
    
    return df_cleaned


def preprocess_old_data(df_old: pd.DataFrame) -> pd.DataFrame:
    """旧マスタを事前処理（文字列変換を1回だけ実行）"""
    df_processed = df_old.copy()
    df_processed['メーカー名称_旧'] = df_processed['メーカー名称_旧'].astype(str).str.strip()
    df_processed['ブランドコード_旧'] = df_processed['ブランドコード_旧'].astype(str).str.strip()
    df_processed['標準分類コード(タイプ)_旧'] = df_processed['標準分類コード(タイプ)_旧'].astype(str).str.strip()
    df_processed['目付_旧_float'] = pd.to_numeric(df_processed['目付_旧'], errors='coerce')
    return df_processed


def process_master_data(old_path: str, new_path: str, output_dir: str):
    """メインロジック：新品→旧品マッチング"""
    
    try:
        print("データ読み込み中...")
        df_new = load_data(new_path, '_新')
        df_old = load_data(old_path, '_旧')
        print(f"【デバッグ】新マスタ読み込み後: {len(df_new)}件")
        print(f"【デバッグ】旧マスタ読み込み後: {len(df_old)}件")
        
        # 初期クレンジング
        df_new = clean_initial_data(df_new, '_新')
        df_old = clean_initial_data(df_old, '_旧')
        print(f"【デバッグ】新マスタクレンジング後: {len(df_new)}件")
        print(f"【デバッグ】旧マスタクレンジング後: {len(df_old)}件")
        
        if '発売日_旧' not in df_old.columns:
            df_old['発売日_旧'] = pd.NA
        
        # 旧マスタを事前処理（高速化）
        print("旧マスタ前処理中...")
        df_old_processed = preprocess_old_data(df_old)
        
        print(f"突合処理開始...（{len(df_new)}件）")
        
        # 結果を格納するリスト
        results = []
        
        for idx, new_row in df_new.iterrows():
            if idx % 100 == 0:
                print(f"処理中... {idx}/{len(df_new)}")
            
            result = find_best_match(new_row, df_old_processed, df_old)
            results.append(result)
        
        print(f"【デバッグ】ループ処理結果: {len(results)}件")
        
        # 結果をDataFrameに変換
        print("結果を整形中...")
        analysis_result = pd.DataFrame(results)
        print(f"【デバッグ】analysis_result: {len(analysis_result)}件")
        
        final_df = pd.concat([df_new.reset_index(drop=True), analysis_result], axis=1)
        print(f"【デバッグ】最終DataFrame: {len(final_df)}件")
        
        final_df = final_df.rename(columns={
            '標準分類名(クラス)_新': '標準分類(クラス)_新',
        }, errors='ignore')
        
        report_columns = [
            '照合結果', 'JANコード_旧','商品名称（カナ）_旧','メーカー名称_旧','標準分類(クラス)_旧', 
            'ブランド名称_旧','目付_旧','発売日_旧','最高類似度','判定',
            'JANコード_新','商品名称（カナ）_新','メーカー名称_新','標準分類(クラス)_新', 
            'ブランド名称_新','目付_新',
            '候補', 'スキップ理由',
        ]
        
        existing_columns = [col for col in report_columns if col in final_df.columns]
        output_df = final_df[existing_columns].copy()
        output_df = output_df.fillna('')
        
        # CSV出力
        print("CSV出力中...")
        csv_path = Path(output_dir) / 'マッチング結果.csv'
        output_df.to_csv(csv_path, index=False, encoding='utf-8')
        
        # 候補品リスト作成（候補ありのみ）
        print("候補品リスト作成中...")
        candidate_sheet_data = create_candidate_sheet_data(df_new, df_old_processed, df_old, results)
        
        # Excel出力
        print("Excel出力中...")
        excel_path = Path(output_dir) / 'マッチング結果.xlsx'
        write_excel_with_dropdowns(output_df, excel_path, candidate_sheet_data)
        
        print("完了！")
        return csv_path, excel_path
    
    except Exception as e:
        raise Exception(f"データ処理中にエラー: {e}")


def find_best_match(new_row: pd.Series, df_old_processed: pd.DataFrame, df_old_original: pd.DataFrame) -> dict:
    """1行の新品に対して最適な旧品を検索（新ロジック）"""
    
    # 新品情報取得
    new_maker_name = str(new_row.get('メーカー名称_新')).strip() if pd.notna(new_row.get('メーカー名称_新')) else None
    new_brand = str(new_row.get('ブランドコード_新')).strip() if pd.notna(new_row.get('ブランドコード_新')) else None
    new_type = str(new_row.get('標準分類コード(タイプ)_新')).strip() if pd.notna(new_row.get('標準分類コード(タイプ)_新')) else None
    new_weight = new_row.get('目付_新')
    new_name = new_row.get('商品名称（カナ）_新')
    
    # nanチェック
    if new_maker_name == 'nan': new_maker_name = None
    if new_brand == 'nan': new_brand = None
    if new_type == 'nan': new_type = None
    
    skip_reasons = []
    matching_old = None
    
    # ========== 新ロジック ==========
    
    # パターン1: ブランドあり
    if new_brand:
        # ブランドコード一致
        matching_old = df_old_processed[df_old_processed['ブランドコード_旧'] == new_brand]
        
        if matching_old.empty:
            return {
                '照合結果': '候補なし（ブランド不一致）',
                '最高類似度': 0.0, '判定': '✕', '候補': '', 'スキップ理由': '', '候補あり': False
            }
        
        # 目付チェック
        if pd.notna(new_weight):
            min_w, max_w = get_weight_range(new_weight)
            if min_w and max_w:
                weight_filtered = matching_old[
                    (matching_old['目付_旧_float'] >= min_w) & 
                    (matching_old['目付_旧_float'] <= max_w)
                ]
                if weight_filtered.empty:
                    return {
                        '照合結果': '候補なし（目付範囲外）',
                        '最高類似度': 0.0, '判定': '✕', '候補': '', 'スキップ理由': '', '候補あり': False
                    }
                matching_old = weight_filtered
        else:
            skip_reasons.append('目付スキップ')
    
    # パターン2: ブランドなし → メーカー名称+タイプ
    elif new_maker_name and new_type:
        # メーカー名称+タイプ一致
        matching_old = df_old_processed[
            (df_old_processed['メーカー名称_旧'] == new_maker_name) &
            (df_old_processed['標準分類コード(タイプ)_旧'] == new_type)
        ]
        
        if matching_old.empty:
            return {
                '照合結果': '候補なし（メーカー名称+タイプ不一致）',
                '最高類似度': 0.0, '判定': '✕', '候補': '', 'スキップ理由': '', '候補あり': False
            }
        
        # 目付チェック
        if pd.notna(new_weight):
            min_w, max_w = get_weight_range(new_weight)
            if min_w and max_w:
                weight_filtered = matching_old[
                    (matching_old['目付_旧_float'] >= min_w) & 
                    (matching_old['目付_旧_float'] <= max_w)
                ]
                if weight_filtered.empty:
                    return {
                        '照合結果': '候補なし（目付範囲外）',
                        '最高類似度': 0.0, '判定': '✕', '候補': '', 'スキップ理由': '', '候補あり': False
                    }
                matching_old = weight_filtered
        else:
            skip_reasons.append('目付スキップ')
    
    else:
        return {
            '照合結果': '候補なし（キーコード不足）',
            '最高類似度': 0.0, '判定': '✕', '候補': '', 'スキップ理由': '', '候補あり': False
        }
    
    # ========== 名称一致で最高類似度を選択 ==========
    candidates = matching_old[['商品名称（カナ）_旧', 'JANコード_旧']].drop_duplicates()
    
    if candidates.empty:
        return {
            '照合結果': '候補なし（名称一致なし）',
            '最高類似度': 0.0, '判定': '✕', '候補': '',
            'スキップ理由': '、'.join(skip_reasons) if skip_reasons else '', '候補あり': False
        }
    
    similarities = [
        (calculate_similarity(new_name, row['商品名称（カナ）_旧']), 
         row['商品名称（カナ）_旧'], 
         row['JANコード_旧'])
        for _, row in candidates.iterrows()
    ]
    similarities.sort(key=lambda x: x[0], reverse=True)
    
    best_score, best_name, best_jan = similarities[0]
    best_old_row = df_old_original[df_old_original['JANコード_旧'] == best_jan].iloc[0]
    
    # 判定
    if best_score >= 0.8:
        result = '高類似度候補あり (80%以上)'
        judgment = '○'
    else:
        result = '低類似度 (80%未満・要手動確認)'
        judgment = '✕'
    
    return {
        '照合結果': result,
        '最高類似度': best_score,
        '判定': judgment,
        '候補': f"{best_name}({best_score:.1%})",
        'スキップ理由': '、'.join(skip_reasons) if skip_reasons else '',
        '候補あり': True,
        'JANコード_旧': best_old_row.get('JANコード_旧', ''),
        '商品名称（カナ）_旧': best_old_row.get('商品名称（カナ）_旧', ''),
        'メーカー名称_旧': best_old_row.get('メーカー名称_旧', ''),
        '標準分類(クラス)_旧': best_old_row.get('標準分類名(クラス)_旧', ''),
        'ブランド名称_旧': best_old_row.get('ブランド名称_旧', ''),
        '目付_旧': best_old_row.get('目付_旧', ''),
        '発売日_旧': best_old_row.get('発売日_旧', ''),
    }


def create_candidate_sheet_data(df_new, df_old_processed, df_old_original, results) -> list:
    """候補品リスト作成（候補ありのみ）"""
    candidates_list = []
    
    for idx, new_row in df_new.iterrows():
        if idx % 100 == 0:
            print(f"候補品処理中... {idx}/{len(df_new)}")
        
        # 候補なしはスキップ
        if not results[idx].get('候補あり', False):
            continue
        
        new_maker_name = str(new_row.get('メーカー名称_新')).strip() if pd.notna(new_row.get('メーカー名称_新')) else None
        new_brand = str(new_row.get('ブランドコード_新')).strip() if pd.notna(new_row.get('ブランドコード_新')) else None
        new_type = str(new_row.get('標準分類コード(タイプ)_新')).strip() if pd.notna(new_row.get('標準分類コード(タイプ)_新')) else None
        new_name = new_row.get('商品名称（カナ）_新', '')
        new_weight = new_row.get('目付_新')
        
        if new_maker_name == 'nan': new_maker_name = None
        if new_brand == 'nan': new_brand = None
        if new_type == 'nan': new_type = None
        
        new_jan = new_row.get('JANコード_新', '')
        new_maker_name_val = new_row.get('メーカー名称_新', '')
        new_class = new_row.get('標準分類(クラス)_新', '')
        new_brand_name = new_row.get('ブランド名称_新', '')
        
        all_matches = []
        
        # パターンA: ブランド+目付
        if new_brand and pd.notna(new_weight):
            min_w, max_w = get_weight_range(new_weight)
            if min_w and max_w:
                pattern_a = df_old_processed[
                    (df_old_processed['ブランドコード_旧'] == new_brand) &
                    (df_old_processed['目付_旧_float'] >= min_w) &
                    (df_old_processed['目付_旧_float'] <= max_w)
                ].copy()
                pattern_a['パターン'] = 'ブランド+目付'
                all_matches.append(pattern_a)
        
        # パターンB: ブランドのみ
        if new_brand:
            pattern_b = df_old_processed[
                df_old_processed['ブランドコード_旧'] == new_brand
            ].copy()
            pattern_b['パターン'] = 'ブランドのみ'
            all_matches.append(pattern_b)
        
        # パターンC: メーカー名称+タイプ+目付
        if new_maker_name and new_type and pd.notna(new_weight):
            min_w, max_w = get_weight_range(new_weight)
            if min_w and max_w:
                pattern_c = df_old_processed[
                    (df_old_processed['メーカー名称_旧'] == new_maker_name) &
                    (df_old_processed['標準分類コード(タイプ)_旧'] == new_type) &
                    (df_old_processed['目付_旧_float'] >= min_w) &
                    (df_old_processed['目付_旧_float'] <= max_w)
                ].copy()
                pattern_c['パターン'] = 'メーカー名称+タイプ+目付'
                all_matches.append(pattern_c)
        
        # パターンD: メーカー名称+タイプのみ（新規追加）
        if new_maker_name and new_type:
            pattern_d = df_old_processed[
                (df_old_processed['メーカー名称_旧'] == new_maker_name) &
                (df_old_processed['標準分類コード(タイプ)_旧'] == new_type)
            ].copy()
            pattern_d['パターン'] = 'メーカー名称+タイプのみ'
            all_matches.append(pattern_d)
        
        if all_matches:
            combined = pd.concat(all_matches, ignore_index=True)
            combined = combined.drop_duplicates(subset=['JANコード_旧'])
            
            # 類似度計算
            similarities = [
                (calculate_similarity(new_name, row.get('商品名称（カナ）_旧', '')), row)
                for _, row in combined.iterrows()
            ]
            similarities.sort(key=lambda x: x[0], reverse=True)
            
            for score, old_row in similarities:
                candidates_list.append({
                    '旧候補JANコード': old_row.get('JANコード_旧', ''),
                    '旧候補商品名': old_row.get('商品名称（カナ）_旧', ''),
                    '旧メーカー名称': old_row.get('メーカー名称_旧', ''),
                    '旧標準分類名(クラス)': old_row.get('標準分類名(クラス)_旧', ''),
                    '旧ブランド名称': old_row.get('ブランド名称_旧', ''),
                    '旧目付': old_row.get('目付_旧', ''),
                    '旧発売日': old_row.get('発売日_旧', ''),
                    '類似度': f'{score:.1%}',
                    '新JANコード': new_jan,
                    '新商品名': new_name,
                    '新メーカー名称': new_maker_name_val,
                    '新標準分類名(クラス)': new_class,
                    '新ブランド名称': new_brand_name,
                    '新目付': new_weight,
                    'パターン': old_row.get('パターン', ''),
                })
    
    return candidates_list


def write_excel_with_dropdowns(df, output_path, candidate_data=None):
    """Excel出力"""
    wb = Workbook()
    ws = wb.active
    ws.title = "マッチング"
    
    # ヘッダー
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # データ
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # 列幅調整
    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20
    
    # 候補品リスト
    if candidate_data:
        df_candidates = pd.DataFrame(candidate_data).fillna('')
        ws_candidates = wb.create_sheet("候補品リスト")
        
        for col_idx, col_name in enumerate(df_candidates.columns, 1):
            cell = ws_candidates.cell(row=1, column=col_idx, value=col_name)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for row_idx, row in enumerate(dataframe_to_rows(df_candidates, index=False, header=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_candidates.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        for col_idx in range(1, len(df_candidates.columns) + 1):
            ws_candidates.column_dimensions[ws_candidates.cell(row=1, column=col_idx).column_letter].width = 20
    
    wb.save(output_path)


class MasterMatcherApp:
    def __init__(self, master):
        self.master = master
        master.title("旧品マッチング（新→旧方向）")
        master.geometry("400x250")
        
        self.old_path_var = tk.StringVar()
        self.new_path_var = tk.StringVar()
        self.output_dir_var = tk.StringVar(value=os.path.join(os.path.expanduser('~'), 'Desktop'))
        
        self.create_widgets(master)
    
    def create_widgets(self, master):
        main_frame = tk.Frame(master, padx=10, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(main_frame, text="旧マスタファイル (.csv, .tsv, .xlsx)").grid(row=0, column=0, sticky='w', pady=5)
        tk.Entry(main_frame, textvariable=self.old_path_var, width=50).grid(row=1, column=0, padx=5, sticky='ew')
        tk.Button(main_frame, text="参照", command=self.select_file_old).grid(row=1, column=1, padx=5)
        
        tk.Label(main_frame, text="新マスタファイル (.csv, .tsv, .xlsx)").grid(row=2, column=0, sticky='w', pady=5)
        tk.Entry(main_frame, textvariable=self.new_path_var, width=50).grid(row=3, column=0, padx=5, sticky='ew')
        tk.Button(main_frame, text="参照", command=self.select_file_new).grid(row=3, column=1, padx=5)
        
        tk.Label(main_frame, text="結果出力先フォルダ").grid(row=4, column=0, sticky='w', pady=5)
        tk.Entry(main_frame, textvariable=self.output_dir_var, width=50).grid(row=5, column=0, padx=5, sticky='ew')
        tk.Button(main_frame, text="参照", command=self.select_output_dir).grid(row=5, column=1, padx=5)
        
        tk.Button(main_frame, text="マッチング開始", command=self.execute_analysis,
                  bg="blue", fg="white", font=('Helvetica', 12, 'bold')).grid(row=6, column=0, columnspan=2, pady=20, sticky='ew')
    
    def select_file(self, path_var):
        file_types = [
            ("マスタファイル", "*.csv *.tsv *.xlsx *.xls"),
            ("CSVファイル", "*.csv"),
            ("TSVファイル", "*.tsv"),
            ("Excelファイル", "*.xlsx *.xls")
        ]
        file_path = filedialog.askopenfilename(parent=self.master, title="ファイルを選択", filetypes=file_types)
        if file_path:
            path_var.set(file_path)
    
    def select_file_old(self):
        self.select_file(self.old_path_var)
    
    def select_file_new(self):
        self.select_file(self.new_path_var)
    
    def select_output_dir(self):
        folder_path = filedialog.askdirectory(parent=self.master, title="保存フォルダを選択")
        if folder_path:
            self.output_dir_var.set(folder_path)
    
    def execute_analysis(self):
        new_path = self.new_path_var.get()
        old_path = self.old_path_var.get()
        output_dir = self.output_dir_var.get()
        
        if not all([new_path, old_path, output_dir]):
            messagebox.showwarning("入力不足", "すべて選択してください")
            return
        
        try:
            csv_path, excel_path = process_master_data(old_path, new_path, output_dir)
            messagebox.showinfo("完了", f"✅ 完了\n\nCSV: {csv_path}\nExcel: {excel_path}")
        except Exception as e:
            messagebox.showerror("エラー", f"❌ エラー発生\n{e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = MasterMatcherApp(root)
    root.mainloop()